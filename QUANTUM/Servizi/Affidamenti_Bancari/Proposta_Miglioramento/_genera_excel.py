import sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Palette colori Quantum
ORO       = "B8962E"
ORO_LIGHT = "F5EDD5"
NERO      = "1A1A1A"
GRIGIO_H  = "F0EDE6"
GRIGIO_R  = "FAFAF7"
BIANCO    = "FFFFFF"
ROSSO_W   = "C00000"

# Anno corrente e precedente
A1, A2 = "2024", "2025"

def oro_fill():   return PatternFill("solid", fgColor=ORO)
def light_fill(): return PatternFill("solid", fgColor=ORO_LIGHT)
def grey_fill():  return PatternFill("solid", fgColor=GRIGIO_H)
def row_fill():   return PatternFill("solid", fgColor=GRIGIO_R)

def bold(size=10, color=BIANCO, italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)
def normal(size=10, color=NERO, bold=False):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def medium_border():
    s = Side(style='medium', color=ORO)
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def left(wrap=False):   return Alignment(horizontal='left',   vertical='center', wrap_text=wrap)

def header_row(ws, row, cols, titles, heights=None):
    """Scrive una riga di intestazione colonne con stile oro."""
    for i, (col, title) in enumerate(zip(cols, titles)):
        c = ws.cell(row=row, column=col, value=title)
        c.font      = bold(9)
        c.fill      = oro_fill()
        c.alignment = center(wrap=True)
        c.border    = thin_border()

def sub_header(ws, row, col, value, span_end=None, fg=GRIGIO_H):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = bold(9, color=NERO)
    c.fill      = PatternFill("solid", fgColor=fg)
    c.alignment = center(wrap=True)
    c.border    = thin_border()
    if span_end:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=span_end)

def intestazione_foglio(ws, titolo, max_col):
    """Riga 1: titolo Quantum — riga 2: ragione sociale + data."""
    # Riga 1 — titolo foglio
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=f"QUANTUM S.r.l.  |  {titolo}")
    c.font      = Font(name="Calibri", bold=True, size=13, color=BIANCO)
    c.fill      = PatternFill("solid", fgColor=NERO)
    c.alignment = left()

    # Riga 2 — Ragione Sociale / Data
    ws.row_dimensions[2].height = 22
    ws.cell(row=2, column=1, value="Ragione Sociale:").font = bold(9, color=NERO)
    ws.cell(row=2, column=1).fill = light_fill()
    ws.cell(row=2, column=1).alignment = left()
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=max_col - 3)
    c2 = ws.cell(row=2, column=2, value="")
    c2.fill   = PatternFill("solid", fgColor="FFFEF8")
    c2.border = Border(bottom=Side(style='medium', color=ORO))
    c2.alignment = left()

    ws.cell(row=2, column=max_col - 2, value="Data compilazione:").font = bold(9, color=NERO)
    ws.cell(row=2, column=max_col - 2).fill = light_fill()
    ws.cell(row=2, column=max_col - 2).alignment = center()
    ws.merge_cells(start_row=2, start_column=max_col - 1, end_row=2, end_column=max_col)
    c3 = ws.cell(row=2, column=max_col - 1, value="")
    c3.border = Border(bottom=Side(style='medium', color=ORO))
    c3.alignment = center()

    # Riga 3 — separatore vuoto sottile
    ws.row_dimensions[3].height = 6

def data_row(ws, row, col_start, col_end, shade=False):
    """Applica stile riga dati."""
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        if shade:
            c.fill = row_fill()
        c.border = thin_border()
        c.alignment = center()

def nota_riga(ws, row, col, testo, max_col):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=col, value=testo)
    c.font      = Font(name="Calibri", italic=True, size=8, color="777777")
    c.alignment = left()

def riga_totale(ws, row, col_label, label, col_val, formula_range, max_col, num_format="#,##0.000"):
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=col_label, end_row=row, end_column=col_val - 1)
    c = ws.cell(row=row, column=col_label, value=label)
    c.font      = bold(9, color=BIANCO)
    c.fill      = PatternFill("solid", fgColor=ORO)
    c.alignment = left()
    c.border    = thin_border()

    ct = ws.cell(row=row, column=col_val, value=f"=SUM({formula_range})")
    ct.font         = bold(9, color=BIANCO)
    ct.fill         = PatternFill("solid", fgColor=ORO)
    ct.alignment    = center()
    ct.border       = thin_border()
    ct.number_format = num_format
    ws.merge_cells(start_row=row, start_column=col_val, end_row=row, end_column=max_col)

# ════════════════════════════════════════════════════════════
wb = Workbook()

# ── 0. ISTRUZIONI ────────────────────────────────────────────
ws0 = wb.active
ws0.title = "ISTRUZIONI"
ws0.sheet_properties.tabColor = NERO
ws0.column_dimensions['A'].width = 90
ws0.row_dimensions[1].height = 30
ws0.merge_cells("A1:A1")
c = ws0.cell(row=1, column=1, value="QUANTUM S.r.l.  |  Dettaglio Affidamenti e Debiti Tributari — GUIDA ALLA COMPILAZIONE")
c.font      = Font(name="Calibri", bold=True, size=13, color=BIANCO)
c.fill      = PatternFill("solid", fgColor=NERO)
c.alignment = left()

istruzioni = [
    ("", ""),
    ("FOGLIO", "COSA COMPILARE"),
    ("MEDIOLUNGO", "Finanziamenti e prestiti a medio-lungo termine in essere (mutui, prestiti chirografari, obbligazioni). Includere anche i finanziamenti in fase di richiesta."),
    ("BREVE", "Affidamenti bancari a breve termine (fidi in c/c, conto anticipi, SBF, castelletti). Includere anche quelli in fase di richiesta."),
    ("LEASING", "Contratti di leasing in essere (auto, macchinari, immobili). Inserire anche i canoni totali a conto economico degli ultimi due esercizi."),
    ("TRIBUTARI/PREVID.", "Debiti tributari o previdenziali rateizzati con Agenzia delle Entrate, Agenzia della Riscossione o INPS."),
    ("", ""),
    ("NOTE GENERALI", ""),
    ("Importi", "Tutti gli importi in Euro/migliaia (es. 150.000€ = 150,000)."),
    ("Date", "Formato GG/MM/AAAA."),
    ("Righe vuote", "Lasciare vuote le righe non utilizzate. Non eliminare le righe."),
    ("Totali", "Le righe TOTALE si calcolano automaticamente — non modificarle."),
    ("Colori", "Sfondo oro = intestazione. Sfondo giallo chiaro = cella da compilare. Sfondo grigio = riga alternata per leggibilità."),
    ("", ""),
    ("CONTATTI", "Per assistenza: Quantum S.r.l. — Via Castelletta 8, 63831 Rapagnano (FM) — Tel. +39 0734 63831 — quantum@pec.cloud"),
]

for i, (foglio, desc) in enumerate(istruzioni, start=3):
    ws0.row_dimensions[i].height = 18
    ca = ws0.cell(row=i, column=1)
    if foglio == "FOGLIO":
        ca.value = f"{'FOGLIO':<20}{'DESCRIZIONE'}"
        ca.font  = bold(10, color=BIANCO)
        ca.fill  = oro_fill()
        ca.alignment = left()
    elif foglio in ("NOTE GENERALI", "CONTATTI"):
        ca.value = foglio
        ca.font  = bold(10, color=NERO)
        ca.fill  = light_fill()
        ca.alignment = left()
    elif foglio:
        ca.value = f"{'  › ' + foglio:<22}{desc}"
        ca.font  = normal(10)
        ca.alignment = left(wrap=True)
        ws0.row_dimensions[i].height = 30 if len(desc) > 80 else 18
    else:
        ca.value = ""

# ── 1. MEDIOLUNGO ────────────────────────────────────────────
ws1 = wb.create_sheet("MEDIOLUNGO")
ws1.sheet_properties.tabColor = ORO
MAX1 = 11

# Larghezze colonne
col_w1 = [22, 18, 12, 12, 12, 14, 16, 14, 10, 28, 14]
for i, w in enumerate(col_w1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

intestazione_foglio(ws1, "Finanziamenti e Prestiti a Medio-Lungo Termine (Euro/migliaia)", MAX1)

# Riga 4 — titolo tabella
ws1.row_dimensions[4].height = 20
ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=MAX1)
c = ws1.cell(row=4, column=1,
    value=f"Finanziamenti / Prestiti Obbligazionari in essere alla data di compilazione — compresi quelli in richiesta")
c.font      = bold(9, color=NERO)
c.fill      = light_fill()
c.alignment = left(wrap=True)
c.border    = thin_border()

# Riga 5 — intestazioni colonne
ws1.row_dimensions[5].height = 36
header_row(ws1, 5,
    list(range(1, MAX1 + 1)),
    ["Istituto / Banca", "Tipo Finanziamento (*)", "Data Erogazione",
     "Data Inizio", "Data Scadenza", "Importo Erogato",
     "Residuo Mese Prec.", "Importo Rata", "Moratoria\n(SI/NO)",
     "Scadenza 1ª Rata Capitale\n(post pre-amm. / moratoria)", "Periodicità Rata"]
)

# Righe dati (10 righe)
DATA_START1 = 6
for r in range(DATA_START1, DATA_START1 + 10):
    ws1.row_dimensions[r].height = 18
    shade = (r % 2 == 0)
    data_row(ws1, r, 1, MAX1, shade)
    for col in [6, 7, 8]:
        ws1.cell(row=r, column=col).number_format = "#,##0.000"

# Riga totale
TOT1 = DATA_START1 + 10
ws1.row_dimensions[TOT1].height = 20
riga_totale(ws1, TOT1, 1, "TOTALE RESIDUO", 7,
            f"G{DATA_START1}:G{TOT1-1}", MAX1)

# Nota
nota_riga(ws1, TOT1 + 1, 1,
    "(*) Tipologie: Mutuo ipotecario / Prestito chirografario / Prestito obbligazionario / Finanziamento agevolato / Altro",
    MAX1)

# ── 2. BREVE ─────────────────────────────────────────────────
ws2 = wb.create_sheet("BREVE")
ws2.sheet_properties.tabColor = ORO
MAX2 = 7

col_w2 = [24, 20, 28, 16, 16, 16, 18]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

intestazione_foglio(ws2, "Affidamenti a Breve Termine (Euro/migliaia)", MAX2)

ws2.row_dimensions[4].height = 20
ws2.merge_cells(start_row=4, start_column=1, end_row=4, end_column=MAX2)
c = ws2.cell(row=4, column=1,
    value="Affidamenti bancari a breve termine in essere alla data di compilazione — compresi quelli in richiesta")
c.font = bold(9, color=NERO); c.fill = light_fill()
c.alignment = left(wrap=True); c.border = thin_border()

ws2.row_dimensions[5].height = 36
header_row(ws2, 5, list(range(1, MAX2 + 1)),
    ["Istituto / Banca", "Riferimento / Numero Pratica",
     "Tipologia Affidamento (*)", "Importo Concesso",
     "Importo Utilizzato", "Importo Residuo / Disponibile", "Scadenza Linea"]
)

DATA_START2 = 6
for r in range(DATA_START2, DATA_START2 + 10):
    ws2.row_dimensions[r].height = 18
    shade = (r % 2 == 0)
    data_row(ws2, r, 1, MAX2, shade)
    for col in [4, 5, 6]:
        ws2.cell(row=r, column=col).number_format = "#,##0.000"

TOT2 = DATA_START2 + 10
ws2.row_dimensions[TOT2].height = 20
# Label "TOTALI" su colonne 1-3
ws2.merge_cells(start_row=TOT2, start_column=1, end_row=TOT2, end_column=3)
c_lab = ws2.cell(row=TOT2, column=1, value="TOTALI")
c_lab.font = bold(9, color=BIANCO); c_lab.fill = oro_fill()
c_lab.alignment = left(); c_lab.border = thin_border()
# Totale concesso col 4
c4 = ws2.cell(row=TOT2, column=4, value=f"=SUM(D{DATA_START2}:D{TOT2-1})")
c4.font = bold(9, color=BIANCO); c4.fill = oro_fill()
c4.alignment = center(); c4.border = thin_border()
c4.number_format = "#,##0.000"
# Totale utilizzato col 5
c5 = ws2.cell(row=TOT2, column=5, value=f"=SUM(E{DATA_START2}:E{TOT2-1})")
c5.font = bold(9, color=BIANCO); c5.fill = oro_fill()
c5.alignment = center(); c5.border = thin_border()
c5.number_format = "#,##0.000"
# Totale residuo col 6
c6 = ws2.cell(row=TOT2, column=6, value=f"=SUM(F{DATA_START2}:F{TOT2-1})")
c6.font = bold(9, color=BIANCO); c6.fill = oro_fill()
c6.alignment = center(); c6.border = thin_border()
c6.number_format = "#,##0.000"
# Scadenza col 7 vuota ma stilata
c7 = ws2.cell(row=TOT2, column=7)
c7.fill = oro_fill(); c7.border = thin_border()

nota_riga(ws2, TOT2 + 1, 1,
    "(*) Tipologie: Fido in c/c / Conto anticipi / Castelletto SBF / Anticipo fatture / Factoring / Fideiussione bancaria / Altro",
    MAX2)

# ── 3. LEASING ───────────────────────────────────────────────
ws3 = wb.create_sheet("LEASING")
ws3.sheet_properties.tabColor = ORO
MAX3 = 9

col_w3 = [26, 20, 14, 14, 14, 22, 28, 14, 18]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

intestazione_foglio(ws3, "Contratti di Leasing in essere (Euro/migliaia)", MAX3)

# Canoni a conto economico
ws3.row_dimensions[4].height = 20
ws3.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
c = ws3.cell(row=4, column=1, value="Canoni Leasing Totali a Conto Economico")
c.font = bold(9, color=NERO); c.fill = light_fill()
c.alignment = left(); c.border = thin_border()
for col, anno in [(3, A1), (4, A2)]:
    ch = ws3.cell(row=4, column=col, value=anno)
    ch.font = bold(9, color=BIANCO); ch.fill = oro_fill()
    ch.alignment = center(); ch.border = thin_border()
for col in [3, 4]:
    cd = ws3.cell(row=5, column=col, value=0)
    cd.number_format = "#,##0.000"
    cd.alignment = center(); cd.border = thin_border()
    cd.fill = PatternFill("solid", fgColor="FFFEF8")
ws3.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
ws3.cell(row=5, column=1).fill = PatternFill("solid", fgColor="FFFEF8")
ws3.cell(row=5, column=1).border = thin_border()

# Separatore
ws3.row_dimensions[6].height = 8

# Titolo tabella
ws3.row_dimensions[7].height = 20
ws3.merge_cells(start_row=7, start_column=1, end_row=7, end_column=MAX3)
c = ws3.cell(row=7, column=1,
    value=f"Contratti di leasing in essere alla data di compilazione — compresi quelli ottenuti nell'esercizio {A2}")
c.font = bold(9, color=NERO); c.fill = light_fill()
c.alignment = left(wrap=True); c.border = thin_border()

ws3.row_dimensions[8].height = 36
header_row(ws3, 8, list(range(1, MAX3 + 1)),
    ["Società di Leasing", "Oggetto del Leasing (*)",
     "Data Inizio", "Data Scadenza", "Canone Annuo",
     "Residuo Mese Precedente",
     "Scadenza 1ª Rata Capitale\n(post pre-amm. / moratoria)",
     "Importo Rata", "Periodicità Rata"]
)

DATA_START3 = 9
for r in range(DATA_START3, DATA_START3 + 8):
    ws3.row_dimensions[r].height = 18
    shade = (r % 2 == 0)
    data_row(ws3, r, 1, MAX3, shade)
    for col in [5, 6, 8]:
        ws3.cell(row=r, column=col).number_format = "#,##0.000"

TOT3 = DATA_START3 + 8
riga_totale(ws3, TOT3, 1, "TOTALE RESIDUO LEASING", 6,
            f"F{DATA_START3}:F{TOT3-1}", MAX3)

nota_riga(ws3, TOT3 + 1, 1,
    "(*) Oggetto: Leasing immobiliare / Leasing strumentale / Leasing auto-moto / Leasing nautico / Altro",
    MAX3)

# ── 4. TRIBUTARI/PREVIDENZIALI ───────────────────────────────
ws4 = wb.create_sheet("TRIBUTARI-PREVID.")
ws4.sheet_properties.tabColor = ORO
MAX4 = 9

col_w4 = [24, 30, 12, 12, 12, 16, 18, 14, 18]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

intestazione_foglio(ws4, "Debiti Tributari e Previdenziali Rateizzati (Euro/migliaia)", MAX4)

ws4.row_dimensions[4].height = 20
ws4.merge_cells(start_row=4, start_column=1, end_row=4, end_column=MAX4)
c = ws4.cell(row=4, column=1,
    value="Debiti tributari/previdenziali rateizzati in essere alla data di compilazione (*)")
c.font = bold(9, color=NERO); c.fill = light_fill()
c.alignment = left(wrap=True); c.border = thin_border()

ws4.row_dimensions[5].height = 36
header_row(ws4, 5, list(range(1, MAX4 + 1)),
    ["N. Comunicazione / Cartella", "Tipologia Debito Rateizzato (*)",
     "Data Inizio Rateizzazione", "Data Scadenza",
     "Durata\n(mesi)", "Importo Originario",
     "Residuo alla Data (**)", "Importo Rata", "Periodicità Rata"]
)

DATA_START4 = 6
for r in range(DATA_START4, DATA_START4 + 10):
    ws4.row_dimensions[r].height = 18
    shade = (r % 2 == 0)
    data_row(ws4, r, 1, MAX4, shade)
    for col in [6, 7, 8]:
        ws4.cell(row=r, column=col).number_format = "#,##0.000"

TOT4 = DATA_START4 + 10
riga_totale(ws4, TOT4, 1, "TOTALE RESIDUO", 7,
            f"G{DATA_START4}:G{TOT4-1}", MAX4)

nota_riga(ws4, TOT4 + 1, 1,
    "(*) Tipologie: INPS / Agenzia delle Entrate / Agenzia della Riscossione / Comune / Altro ente",
    MAX4)
nota_riga(ws4, TOT4 + 2, 1,
    "(**) Inserire il residuo alla data indicata nel campo 'Data di compilazione' del foglio.",
    MAX4)

# ── Ordine fogli ─────────────────────────────────────────────
wb.move_sheet("ISTRUZIONI", offset=0)

# ── Salva ────────────────────────────────────────────────────
out = r"e:\Lavori Code\QUANTUM\Servizi\Affidamenti_Bancari\Proposta_Miglioramento\Dettaglio_Affidamenti_Quantum.xlsx"
wb.save(out)
print(f"OK — file salvato: {out}")
